from django.db.models.signals import post_save, pre_save, m2m_changed
from django.dispatch import receiver
from django.contrib import messages
import requests
from openpyxl import Workbook, load_workbook
 
from django.conf import settings
 

#local imports
from .models import Documents, Students, Payments, BotMessages, StudentUser_ids


@receiver(post_save, sender=Documents)
def create_document(sender, instance, created, *args, **kwargs):
    if created:
        wb1 = load_workbook(instance.document)
        wb = wb1['IFP']
        ws = wb1.active
        counter = 0
        Payments.objects.all().delete()
        for i in range(4, ws.max_row + 1):
            if ws.cell(i, 2).value =='ЖАМИ':
                break
            if ws.cell(i, 2).value is not None:
                lang = 'en' if ws.cell(i, 8).value=='инглиз' else 'ru'
                student, created = Students.objects.get_or_create(
                    fish=ws.cell(i, 2).value,
                    id_raqam=ws.cell(i, 3).value,
                )
                student.date_contracted=ws.cell(i, 4).value
                student.contract_soums=ws.cell(i, 5).value
                student.level=ws.cell(i, 6).value
                student.faculty=ws.cell(i, 7).value
                student.edu_lang=lang
                student.remains_year_begin=ws.cell(i, 9).value
                student.save()
                counter += 1
                j=i+1
                while ws.cell(j, 2).value == None:
                
                    pay=ws.cell(j, 11).value
                    
                    if str(ws.cell(j, 11).value).startswith('='):
                        payment = ws.cell(j, 11).value
                        payment = payment.replace("=", "").split('+')
                        pay = sum(int(pym) for pym in payment)
                    payment, created = Payments.objects.get_or_create(
                    student=student,
                    date_paid=ws.cell(j, 10).value,
                    soums_paid=int(pay) or ws.cell(j, 11).value
                    )
                    phone_number=ws.cell(i, 15).value
                    if phone_number:
                        StudentUser_ids.objects.get_or_create(student=student, phone_number=phone_number)
                    j+=1
                i=j            
            else:
                continue


@receiver(m2m_changed, sender=BotMessages.students.through)
def bot_message_created(sender, **kwargs):
    instance = kwargs.pop('instance', None)
    students = kwargs.pop('students', None)
    message = kwargs.pop('message', None)
    admin = kwargs.pop('admin', None)
    action = kwargs.pop('action', None)
    instance.save()
    if action == 'post_add':
        counter, all_obj = 0, 0
        for std in instance.students.all():
            for phone in std.phones.all():
                r = requests.get(f'https://api.telegram.org/bot{settings.BOT_TOKEN}/sendMessage',
                params = {'chat_id':phone.user_id, 'text':instance.message}
                )
                if r.status_code == 200:
                    counter += 1
                all_obj += 1
    return instance


# @receiver(post_save, sender=BotMessages)
# def send_admin_message_to_students(sender, created, instance, *args, **kwargs):
#     all_obj, counter = 0, 0 

    
    
        # messages.add_message(request, messages.WARNING, f'Не отправлено сообщение {all_obj-counter} из {all_obj} студентов')
        # messages.add_message(request, messages.INFO, f'Отправлено сообщение {counter} из {all_obj} студентов')
        
